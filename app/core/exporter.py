"""
ExcelExporter: export RecipeTree instances to an .xlsx workbook.
"""

import os, logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from core.base import EXCEL_COLUMNS


class ExcelExporter:
    """
    Export parameters and formula values into an Excel workbook.
    """

    def export(self, trees, excel_path: str):
        """
        Write out one Excel sheet per RecipeTree, exporting all parameters and formula values.

        Header row is formatted with Arial 10pt bold black on #F2F2F2,
        and panes are frozen below the header, and columns auto-resize to fit content.

        This method accepts either a single RecipeTree or a list of them, then removes the default
        OpenPyXL sheet and prepares a new workbook.  It iterates each tree, calls each node's
        `to_excel_row()` to build a flat dict, and collects any extra columns beyond the fixed
        schema.  After logging how many rows each sheet will have, it writes a header row (fixed
        columns + sorted extras) followed by one row per node.  Finally, it saves the workbook to
        `excel_path` and logs the completion.  Errors during file creation or serialization raise
        the underlying exception for the caller to handle.
        """
        log = logging.getLogger(__name__)
        if not isinstance(trees, list):
            trees = [trees]
        wb = Workbook()
        wb.remove(wb.active)

        all_extras = set()
        sheet_data = {}
        for t in trees:
            rows = []
            for node in t.parameters + t.formula_values:
                row = node.to_excel_row()
                rows.append(row)
                all_extras.update(k for k in row if k not in EXCEL_COLUMNS)
            sheet = os.path.basename(t.filepath)
            sheet_data[sheet] = rows
            log.info("Prepared %d rows for sheet %s", len(rows), sheet)

        extras = sorted(all_extras)
        header = EXCEL_COLUMNS + extras

        # styling objects
        header_font = Font(name="Arial", size=10, bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="F2F2F2")

        for sheet, rows in sheet_data.items():
            ws = wb.create_sheet(sheet)
            # write header
            ws.append(header)
            # apply formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            # freeze panes below header
            ws.freeze_panes = "A2"

            # write data rows
            for row in rows:
                ws.append([row.get(col, "") for col in header])

            # Auto-size columns
            for col_cells in ws.columns:
                max_length = 0
                column_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_path)
        log.info("Excel written to %s", excel_path)
