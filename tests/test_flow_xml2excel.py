# tests/test_flow_xml2excel.py

import os
import pytest
from lxml import etree
from openpyxl import load_workbook

from core.parser import XMLParser
from core.exporter import ExcelExporter
from core.base import NAMESPACE, NSMAP, EXCEL_COLUMNS

SAMPLE_FULL_XML = f"""<?xml version="1.0" encoding="UTF-8"?>
<RecipeElement xmlns="{NAMESPACE}">
  <RecipeElementID>TEST</RecipeElementID>
  <!-- Three Parameters -->
  <Parameter>
    <Name>XFER5_TARGET_CONC_DB</Name>
    <ERPAlias/>
    <PLCReference>1</PLCReference>
    <Real>0</Real>
    <High>9999</High>
    <Low>0</Low>
    <EngineeringUnits/>
    <Scale>false</Scale>
  </Parameter>
  <Parameter>
    <Name>XFER5_TARGET_SOL_CONC</Name>
    <ERPAlias/>
    <PLCReference>1</PLCReference>
    <Real>0</Real>
    <High>9999</High>
    <Low>0</Low>
    <EngineeringUnits/>
    <Scale>false</Scale>
  </Parameter>
  <Parameter>
    <Name>XFER5_TRANSFER_TYPE</Name>
    <ERPAlias/>
    <PLCReference>1</PLCReference>
    <EnumerationSet>N_OPTION</EnumerationSet>
    <EnumerationMember>OPTION_1</EnumerationMember>
  </Parameter>

  <Steps>
    <InitialStep><Name>INITIALSTEP:1</Name></InitialStep>
    <TerminalStep><Name>TERMINALSTEP:1</Name></TerminalStep>

    <!-- Step without FormulaValue -->
    <Step><Name>$NULL:1</Name><StepRecipeID>$NULL</StepRecipeID></Step>

    <!-- Step with two FormulaValues -->
    <Step>
      <Name>ACQ_REL:1</Name>
      <StepRecipeID>ACQ_REL</StepRecipeID>
      <FormulaValue>
        <Name>X_R_END_PHASE_PROMPT</Name>
        <Display>false</Display>
        <Value/>
        <EnumerationSet>N_DISABLEENABLE</EnumerationSet>
        <EnumerationMember>DISABLE</EnumerationMember>
        <FormulaValueLimit Verification="No_Limits">
          <LowLowLowValue>0.</LowLowLowValue>
          <LowLowValue>0.</LowLowValue>
          <LowValue>0.</LowValue>
          <HighValue>0.</HighValue>
          <HighHighValue>0.</HighHighValue>
          <HighHighHighValue>0.</HighHighHighValue>
        </FormulaValueLimit>
      </FormulaValue>
      <FormulaValue>
        <Name>X_R_END_TYPE</Name>
        <Display>false</Display>
        <Defer>AQREL_END_TYPE</Defer>
        <EnumerationSet>N_ENDTYPE</EnumerationSet>
        <EnumerationMember/>
        <FormulaValueLimit Verification="No_Limits">
          <LowLowLowValue>0.</LowLowLowValue>
          <LowLowValue>0.</LowLowValue>
          <LowValue>0.</LowValue>
          <HighValue>0.</HighValue>
          <HighHighValue>0.</HighHighValue>
          <HighHighHighValue>0.</HighHighHighValue>
        </FormulaValueLimit>
      </FormulaValue>
    </Step>
  </Steps>
</RecipeElement>
"""


@pytest.fixture
def sample_pxml(tmp_path):
    p = tmp_path / "TEST.pxml"
    p.write_text(SAMPLE_FULL_XML, encoding="utf-8")
    return str(p)


def test_full_xml2excel_flow(sample_pxml, tmp_path):
    """
    End-to-end xml2excel workflow test.

    This test verifies that a FactoryTalk Batch `.pxml` file containing
    multiple `<Parameter>` and `<FormulaValue>` nodes is correctly parsed
    into the in-memory model and then exported to an Excel workbook.

    First, it uses XMLParser to load and extract the expected number of
    parameters and formula values from the sample XML.  Next, it calls
    ExcelExporter to write these nodes into an `.xlsx` file and reloads
    the workbook to assert that the correct sheet name and header row
    (matching `EXCEL_COLUMNS`) are present.  The test then reads all
    data rows and ensures their count equals the total nodes extracted.

    Finally, it spot-checks specific rows—verifying enumeration values
    and deferred references—confirming that each node's Excel representation
    matches the original XML content.  Any mismatch causes the test to fail,
    guaranteeing that the xml2excel path is reliable and schema-complete.
    """
    # 1) Parse XML (including any children, none here)
    parser = XMLParser()
    trees = parser.parse(sample_pxml)

    # Expect exactly one RecipeTree
    assert isinstance(trees, list) and len(trees) == 1
    tree = trees[0]
    # Check count of extracted nodes
    assert len(tree.parameters) == 3
    assert len(tree.formula_values) == 2

    # 2) Export to Excel
    excel_path = tmp_path / "out.xlsx"
    exporter = ExcelExporter()
    exporter.export(trees, str(excel_path))

    # 3) Read back the workbook
    wb = load_workbook(str(excel_path))
    # Sheet name should match filename
    assert wb.sheetnames == ["TEST.pxml"]

    ws = wb["TEST.pxml"]
    # Header row
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # The first len(EXCEL_COLUMNS) columns must match
    assert header[: len(EXCEL_COLUMNS)] == EXCEL_COLUMNS

    # 4) Collect the data rows
    data = list(ws.iter_rows(min_row=2, values_only=True))
    # should have 5 rows (3 params + 2 fvs)
    assert len(data) == 5

    # 5) Spot-check a couple of rows:
    # Find row for XFER5_TRANSFER_TYPE
    col_fullpath = header.index("FullPath")
    row = next(
        r for r in data if r[col_fullpath] == "TEST/Parameter[XFER5_TRANSFER_TYPE]"
    )
    # EnumerationMember should be OPTION_1
    idx_enum_member = header.index("EnumerationMember")
    assert row[idx_enum_member] == "OPTION_1"

    # Find deferred FV
    row_fv = next(
        r for r in data if r[col_fullpath].endswith("FormulaValue[X_R_END_TYPE]")
    )
    idx_defer = header.index("Defer")
    assert row_fv[idx_defer] == "AQREL_END_TYPE"
    # Verify FormulaValueLimit_HighValue is "0."
    idx_hv = header.index("FormulaValueLimit_HighValue")
    assert row_fv[idx_hv] == "0."

    # Done
    print("✅ Full xml→excel workflow test passed")
